"""Read-only finance reporting and Excel export services."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import extract, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.models.class_registration import ClassRegistration
from app.models.course_class import CourseClass
from app.models.payment import Payment
from app.models.private_tutoring_request import PrivateTutoringRequest
from app.models.subject import Subject
from app.models.teaching_contract import TeachingContract
from app.models.tutor_profile import TutorProfile
from app.models.user_account import UserAccount
from app.schemas.finance import (
    FinanceReportRow,
    FinanceSummary,
    TutorIncomeSummary,
    TutorIncomeTransaction,
)

MONEY_QUANTUM = Decimal("0.01")
ZERO = Decimal("0.00")
RECOGNIZED_STATUSES = {"SUCCEEDED", "REFUNDED"}


def _money(value: Decimal | int | None) -> Decimal:
    return Decimal(value or 0).quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP)


def _safe_text(value: object | None) -> str:
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _excel_number(value: Decimal) -> int | float:
    normalized = _money(value)
    if normalized == normalized.to_integral_value():
        return int(normalized)
    return float(normalized)


async def get_finance_rows(
    db: AsyncSession,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    month: int | None = None,
    year: int | None = None,
    tutor_id: int | None = None,
    target_type: str | None = None,
    class_id: int | None = None,
    contract_id: int | None = None,
    payment_status: str | None = None,
) -> list[FinanceReportRow]:
    student_account = aliased(UserAccount)
    tutor_account = aliased(UserAccount)
    report_date = func.coalesce(Payment.paid_at, Payment.created_at)

    query = (
        select(
            Payment,
            TeachingContract,
            student_account.full_name.label("student_name"),
            TutorProfile.id.label("tutor_id"),
            tutor_account.full_name.label("tutor_name"),
        )
        .join(student_account, Payment.student_account_id == student_account.id)
        .outerjoin(TeachingContract, Payment.contract_id == TeachingContract.id)
        .outerjoin(TutorProfile, TeachingContract.tutor_id == TutorProfile.id)
        .outerjoin(tutor_account, TutorProfile.account_id == tutor_account.id)
        .order_by(report_date.desc(), Payment.id.desc())
    )

    if date_from is not None:
        query = query.where(func.date(report_date) >= date_from)
    if date_to is not None:
        query = query.where(func.date(report_date) <= date_to)
    if month is not None:
        query = query.where(extract("month", report_date) == month)
    if year is not None:
        query = query.where(extract("year", report_date) == year)
    if tutor_id is not None:
        query = query.where(TeachingContract.tutor_id == tutor_id)
    if target_type is not None:
        query = query.where(Payment.target_type == target_type)
    if class_id is not None:
        query = query.where(TeachingContract.class_id == class_id)
    if contract_id is not None:
        query = query.where(Payment.contract_id == contract_id)
    if payment_status == "PENDING_ALL":
        query = query.where(Payment.status.in_(("CREATED", "PENDING", "REFUND_PENDING")))
    elif payment_status is not None:
        query = query.where(Payment.status == payment_status)

    result = await db.execute(query)
    records = result.all()

    private_ids = {
        payment.target_id
        for payment, *_ in records
        if payment.target_type == "PRIVATE_TUTORING_REQUEST"
    }
    registration_ids = {
        payment.target_id
        for payment, *_ in records
        if payment.target_type == "CLASS_REGISTRATION"
    }

    private_targets: dict[int, dict[str, object]] = {}
    if private_ids:
        private_result = await db.execute(
            select(
                PrivateTutoringRequest.id,
                PrivateTutoringRequest.requested_sessions,
                Subject.id,
                Subject.name,
            )
            .join(Subject, PrivateTutoringRequest.subject_id == Subject.id)
            .where(PrivateTutoringRequest.id.in_(private_ids))
        )
        for request_id, sessions, subject_id, subject_name in private_result.all():
            private_targets[request_id] = {
                "target_name": f"Yêu cầu 1-1 ({sessions} buổi)",
                "class_id": None,
                "subject_id": subject_id,
                "subject_name": subject_name,
            }

    class_targets: dict[int, dict[str, object]] = {}
    if registration_ids:
        class_result = await db.execute(
            select(
                ClassRegistration.id,
                CourseClass.id,
                CourseClass.title,
                Subject.id,
                Subject.name,
            )
            .join(CourseClass, ClassRegistration.class_id == CourseClass.id)
            .join(Subject, CourseClass.subject_id == Subject.id)
            .where(ClassRegistration.id.in_(registration_ids))
        )
        for registration_id, course_class_id, title, subject_id, subject_name in class_result.all():
            class_targets[registration_id] = {
                "target_name": title,
                "class_id": course_class_id,
                "subject_id": subject_id,
                "subject_name": subject_name,
            }

    rows: list[FinanceReportRow] = []
    for payment, contract, student_name, resolved_tutor_id, tutor_name in records:
        target = (
            private_targets.get(payment.target_id, {})
            if payment.target_type == "PRIVATE_TUTORING_REQUEST"
            else class_targets.get(payment.target_id, {})
        )

        gross_amount = _money(payment.amount)
        refund_amount = min(_money(payment.refund_amount), gross_amount)
        net_amount = _money(gross_amount - refund_amount)
        center_gross = _money(payment.center_amount_snapshot)
        tutor_gross = _money(payment.tutor_amount_snapshot)
        recognized = payment.status in RECOGNIZED_STATUSES

        if not recognized:
            allocation_status = "NOT_RECOGNIZED"
            center_adjustment = ZERO
            tutor_adjustment = ZERO
            center_net = ZERO
            tutor_net = ZERO
        elif payment.center_amount_snapshot is None or payment.tutor_amount_snapshot is None:
            allocation_status = "MISSING_SNAPSHOT"
            center_adjustment = ZERO
            tutor_adjustment = ZERO
            center_net = ZERO
            tutor_net = ZERO
        elif _money(center_gross + tutor_gross) != gross_amount:
            allocation_status = "SNAPSHOT_MISMATCH"
            center_adjustment = ZERO
            tutor_adjustment = ZERO
            center_net = ZERO
            tutor_net = ZERO
        else:
            allocation_status = "ALLOCATED"
            if gross_amount > ZERO and refund_amount > ZERO:
                center_adjustment = _money(refund_amount * center_gross / gross_amount)
                tutor_adjustment = _money(refund_amount - center_adjustment)
            else:
                center_adjustment = ZERO
                tutor_adjustment = ZERO
            center_net = _money(max(center_gross - center_adjustment, ZERO))
            tutor_net = _money(max(tutor_gross - tutor_adjustment, ZERO))

        rows.append(
            FinanceReportRow(
                payment_id=payment.id,
                created_at=payment.created_at,
                paid_at=payment.paid_at,
                payment_status=payment.status,
                provider=payment.provider or "MOCK",
                sepay_transaction_id=payment.sepay_transaction_id,
                transfer_content=payment.transfer_content,
                student_account_id=payment.student_account_id,
                student_name=student_name,
                tutor_id=resolved_tutor_id,
                tutor_name=tutor_name,
                contract_id=payment.contract_id,
                target_type=payment.target_type,
                target_id=payment.target_id,
                target_name=target.get("target_name"),
                class_id=target.get("class_id"),
                subject_id=target.get("subject_id"),
                subject_name=target.get("subject_name"),
                billing_cycle_label=payment.billing_cycle_label,
                gross_amount=gross_amount,
                refund_amount=refund_amount,
                refund_at=payment.updated_at if refund_amount > ZERO else None,
                refund_reason=payment.refund_reason,
                net_amount=net_amount,
                center_rate=_money(contract.center_rate_snapshot) if contract else None,
                tutor_rate=_money(contract.tutor_rate_snapshot) if contract else None,
                center_gross=center_gross,
                tutor_gross=tutor_gross,
                center_refund_adjustment=center_adjustment,
                tutor_refund_adjustment=tutor_adjustment,
                center_net=center_net,
                tutor_net=tutor_net,
                allocation_status=allocation_status,
            )
        )

    return rows


def summarize_finance(rows: Iterable[FinanceReportRow]) -> FinanceSummary:
    row_list = list(rows)
    recognized = [row for row in row_list if row.payment_status in RECOGNIZED_STATUSES]
    allocated = [row for row in recognized if row.allocation_status == "ALLOCATED"]
    unallocated = [row for row in recognized if row.allocation_status != "ALLOCATED"]
    return FinanceSummary(
        gross_amount=_money(sum((row.gross_amount for row in recognized), ZERO)),
        refund_amount=_money(sum((row.refund_amount for row in recognized), ZERO)),
        net_amount=_money(sum((row.net_amount for row in recognized), ZERO)),
        center_net=_money(sum((row.center_net for row in allocated), ZERO)),
        tutor_net=_money(sum((row.tutor_net for row in allocated), ZERO)),
        unallocated_net=_money(sum((row.net_amount for row in unallocated), ZERO)),
        transaction_count=len(row_list),
        recognized_count=len(recognized),
        missing_snapshot_count=len(unallocated),
    )


def summarize_tutor_income(rows: Iterable[FinanceReportRow], today: date | None = None) -> TutorIncomeSummary:
    today = today or date.today()
    recognized = [
        row
        for row in rows
        if row.payment_status in RECOGNIZED_STATUSES and row.allocation_status == "ALLOCATED"
    ]
    this_month = [
        row
        for row in recognized
        if row.paid_at and row.paid_at.year == today.year and row.paid_at.month == today.month
    ]
    this_year = [row for row in recognized if row.paid_at and row.paid_at.year == today.year]
    return TutorIncomeSummary(
        this_month=_money(sum((row.tutor_net for row in this_month), ZERO)),
        this_year=_money(sum((row.tutor_net for row in this_year), ZERO)),
        total=_money(sum((row.tutor_net for row in recognized), ZERO)),
        refund_adjustment=_money(
            sum((row.tutor_refund_adjustment for row in recognized), ZERO)
        ),
        transaction_count=len(recognized),
    )


def to_tutor_transactions(rows: Iterable[FinanceReportRow]) -> list[TutorIncomeTransaction]:
    return [
        TutorIncomeTransaction(
            payment_id=row.payment_id,
            paid_at=row.paid_at,
            refund_at=row.refund_at,
            target_type=row.target_type,
            target_name=row.target_name,
            subject_name=row.subject_name,
            billing_cycle_label=row.billing_cycle_label,
            payment_status=row.payment_status,
            gross_amount=row.gross_amount,
            refund_amount=row.refund_amount,
            tutor_rate=row.tutor_rate,
            tutor_gross=row.tutor_gross,
            tutor_refund_adjustment=row.tutor_refund_adjustment,
            tutor_net=row.tutor_net,
        )
        for row in rows
        if row.payment_status in RECOGNIZED_STATUSES and row.allocation_status == "ALLOCATED"
    ]


def build_finance_excel(rows: Iterable[FinanceReportRow]) -> bytes:
    row_list = list(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Chi tiet doanh thu"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    columns = [
        ("Payment ID", "payment_id"),
        ("Ngày tạo", "created_at"),
        ("Ngày thanh toán", "paid_at"),
        ("Trạng thái", "payment_status"),
        ("Nhà cung cấp", "provider"),
        ("Mã GD SePay", "sepay_transaction_id"),
        ("Nội dung CK", "transfer_content"),
        ("Student ID", "student_account_id"),
        ("Học viên", "student_name"),
        ("Tutor ID", "tutor_id"),
        ("Gia sư", "tutor_name"),
        ("Contract ID", "contract_id"),
        ("Loại nguồn", "target_type"),
        ("Target ID", "target_id"),
        ("Tên lớp/yêu cầu", "target_name"),
        ("Class ID", "class_id"),
        ("Subject ID", "subject_id"),
        ("Môn học", "subject_name"),
        ("Chu kỳ", "billing_cycle_label"),
        ("Tổng tiền", "gross_amount"),
        ("Hoàn tiền", "refund_amount"),
        ("Ngày hoàn tiền", "refund_at"),
        ("Lý do hoàn", "refund_reason"),
        ("Doanh thu net", "net_amount"),
        ("Tỷ lệ trung tâm (%)", "center_rate"),
        ("Tỷ lệ gia sư (%)", "tutor_rate"),
        ("Trung tâm gross", "center_gross"),
        ("Gia sư gross", "tutor_gross"),
        ("Điều chỉnh trung tâm", "center_refund_adjustment"),
        ("Điều chỉnh gia sư", "tutor_refund_adjustment"),
        ("Trung tâm net", "center_net"),
        ("Gia sư net", "tutor_net"),
        ("Trạng thái phân bổ", "allocation_status"),
    ]
    sheet.append([title for title, _ in columns])

    date_fields = {"created_at", "paid_at", "refund_at"}
    money_fields = {
        "gross_amount",
        "refund_amount",
        "net_amount",
        "center_gross",
        "tutor_gross",
        "center_refund_adjustment",
        "tutor_refund_adjustment",
        "center_net",
        "tutor_net",
    }
    numeric_fields = {"center_rate", "tutor_rate"}

    for row in row_list:
        values: list[object] = []
        for _, field in columns:
            value = getattr(row, field)
            if field in money_fields or field in numeric_fields:
                values.append(None if value is None else _excel_number(value))
            elif field in date_fields:
                values.append(value)
            elif isinstance(value, str) or value is None:
                values.append(_safe_text(value))
            else:
                values.append(value)
        sheet.append(values)

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="D1D5DB"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")

    field_to_column = {field: index + 1 for index, (_, field) in enumerate(columns)}
    for field in date_fields:
        for cell in sheet.iter_cols(
            min_col=field_to_column[field],
            max_col=field_to_column[field],
            min_row=2,
        ):
            for item in cell:
                item.number_format = "dd/mm/yyyy hh:mm"
    for field in money_fields:
        for cell in sheet.iter_cols(
            min_col=field_to_column[field],
            max_col=field_to_column[field],
            min_row=2,
        ):
            for item in cell:
                item.number_format = '#,##0;[Red](#,##0);-'
    for field in numeric_fields:
        for cell in sheet.iter_cols(
            min_col=field_to_column[field],
            max_col=field_to_column[field],
            min_row=2,
        ):
            for item in cell:
                item.number_format = "0.00"

    widths = {
        "B": 18, "C": 18, "F": 18, "G": 20, "I": 24, "K": 24,
        "O": 28, "R": 18, "S": 16, "V": 18, "W": 28, "AG": 22,
    }
    for index in range(1, len(columns) + 1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(letter, 15)

    last_column = get_column_letter(len(columns))
    if row_list:
        table = Table(
            displayName="FinanceReport",
            ref=f"A1:{last_column}{len(row_list) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.auto_filter.ref = f"A1:{last_column}1"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
