from datetime import datetime
from decimal import Decimal
from io import BytesIO

import pytest
import pytest_asyncio
from fastapi import Depends, HTTPException
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import BigInteger, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import selectinload

from app.core.deps import get_current_user, get_db
from app.main import app
from app.models import Base
from app.models.audit_log import AuditLog
from app.models.payment import Payment
from app.models.private_tutoring_request import PrivateTutoringRequest
from app.models.subject import Subject
from app.models.teaching_contract import TeachingContract
from app.models.tutor_profile import TutorProfile
from app.models.user_account import UserAccount
from app.services.finance import build_finance_excel, get_finance_rows, summarize_finance


@compiles(BigInteger, "sqlite")
def compile_big_int_sqlite(element, compiler, **kw):
    del element, compiler, kw
    return "INTEGER"


engine = create_async_engine(
    "sqlite+aiosqlite:///:memory:",
    connect_args={"check_same_thread": False},
)
TestingSessionLocal = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
current_test_user_id: int | None = None


async def override_get_db():
    async with TestingSessionLocal() as session:
        yield session


async def override_get_current_user(db: AsyncSession = Depends(get_db)):
    if current_test_user_id is None:
        raise HTTPException(status_code=401, detail="Unauthorized")
    result = await db.execute(
        select(UserAccount)
        .options(selectinload(UserAccount.tutor_profile))
        .where(UserAccount.id == current_test_user_id)
    )
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return user


@pytest.fixture(autouse=True)
def manage_dependency_overrides():
    old_overrides = dict(app.dependency_overrides)
    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[get_current_user] = override_get_current_user
    yield
    app.dependency_overrides.clear()
    app.dependency_overrides.update(old_overrides)


@pytest_asyncio.fixture(autouse=True)
async def setup_db():
    global current_test_user_id
    current_test_user_id = None
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    yield
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.drop_all)


@pytest_asyncio.fixture
async def db_session():
    async with TestingSessionLocal() as session:
        yield session


@pytest_asyncio.fixture
async def client():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as api_client:
        yield api_client


async def seed_finance_data(db: AsyncSession):
    staff = UserAccount(
        email="staff@lumin.test",
        password_hash="hash",
        role="STAFF",
        full_name="Nhân viên tài chính",
        status="ACTIVE",
    )
    student = UserAccount(
        email="student-finance@lumin.test",
        password_hash="hash",
        role="STUDENT",
        full_name="=Học viên kiểm thử",
        status="ACTIVE",
    )
    tutor_account = UserAccount(
        email="tutor-finance@lumin.test",
        password_hash="hash",
        role="TUTOR",
        full_name="Gia sư A",
        status="ACTIVE",
    )
    other_tutor_account = UserAccount(
        email="tutor-other@lumin.test",
        password_hash="hash",
        role="TUTOR",
        full_name="Gia sư B",
        status="ACTIVE",
    )
    db.add_all([staff, student, tutor_account, other_tutor_account])
    await db.flush()

    tutor = TutorProfile(account_id=tutor_account.id, verification_status="VERIFIED")
    other_tutor = TutorProfile(
        account_id=other_tutor_account.id,
        verification_status="VERIFIED",
    )
    subject = Subject(name="Toán", status="ACTIVE")
    db.add_all([tutor, other_tutor, subject])
    await db.flush()

    request = PrivateTutoringRequest(
        student_account_id=student.id,
        tutor_id=tutor.id,
        subject_id=subject.id,
        grade_level="Lớp 12",
        goal="Ôn thi",
        requested_sessions=5,
        agreed_fee_per_session=Decimal("200000"),
        mode="ONLINE",
        status="PAID",
    )
    other_request = PrivateTutoringRequest(
        student_account_id=student.id,
        tutor_id=other_tutor.id,
        subject_id=subject.id,
        grade_level="Lớp 10",
        goal="Bổ trợ",
        requested_sessions=3,
        agreed_fee_per_session=Decimal("200000"),
        mode="ONLINE",
        status="PAID",
    )
    db.add_all([request, other_request])
    await db.flush()

    contract = TeachingContract(
        tutor_id=tutor.id,
        private_request_id=request.id,
        commission_name_snapshot="Default Commission",
        center_rate_snapshot=Decimal("30"),
        tutor_rate_snapshot=Decimal("70"),
        status="ACTIVE",
    )
    other_contract = TeachingContract(
        tutor_id=other_tutor.id,
        private_request_id=other_request.id,
        commission_name_snapshot="Default Commission",
        center_rate_snapshot=Decimal("40"),
        tutor_rate_snapshot=Decimal("60"),
        status="ACTIVE",
    )
    editable_contract = TeachingContract(
        tutor_id=tutor.id,
        private_request_id=request.id,
        commission_name_snapshot="Default Commission",
        center_rate_snapshot=Decimal("30"),
        tutor_rate_snapshot=Decimal("70"),
        status="PENDING",
    )
    db.add_all([contract, other_contract, editable_contract])
    await db.flush()

    paid_at = datetime(2026, 6, 15, 10, 0, 0)
    payment = Payment(
        student_account_id=student.id,
        target_type="PRIVATE_TUTORING_REQUEST",
        target_id=request.id,
        contract_id=contract.id,
        amount=Decimal("1000000"),
        status="SUCCEEDED",
        provider="SEPAY",
        paid_at=paid_at,
        center_amount_snapshot=Decimal("300000"),
        tutor_amount_snapshot=Decimal("700000"),
        sepay_transaction_id="TX-001",
    )
    refunded_payment = Payment(
        student_account_id=student.id,
        target_type="PRIVATE_TUTORING_REQUEST",
        target_id=request.id,
        contract_id=contract.id,
        amount=Decimal("1000000"),
        status="REFUNDED",
        provider="SEPAY",
        paid_at=paid_at,
        refund_amount=Decimal("200000"),
        refund_reason="=Điều chỉnh kiểm thử",
        center_amount_snapshot=Decimal("300000"),
        tutor_amount_snapshot=Decimal("700000"),
        sepay_transaction_id="TX-002",
    )
    other_payment = Payment(
        student_account_id=student.id,
        target_type="PRIVATE_TUTORING_REQUEST",
        target_id=other_request.id,
        contract_id=other_contract.id,
        amount=Decimal("500000"),
        status="SUCCEEDED",
        provider="MOCK",
        paid_at=paid_at,
        center_amount_snapshot=Decimal("200000"),
        tutor_amount_snapshot=Decimal("300000"),
    )
    db.add_all([payment, refunded_payment, other_payment])
    await db.commit()
    return {
        "staff": staff,
        "student": student,
        "tutor_account": tutor_account,
        "other_tutor_account": other_tutor_account,
        "tutor": tutor,
        "other_tutor": other_tutor,
        "contract": contract,
        "editable_contract": editable_contract,
    }


@pytest.mark.asyncio
async def test_finance_summary_partial_refund_and_excel(db_session: AsyncSession):
    data = await seed_finance_data(db_session)
    rows = await get_finance_rows(db_session, tutor_id=data["tutor"].id)
    summary = summarize_finance(rows)

    assert len(rows) == 2
    assert summary.gross_amount == Decimal("2000000.00")
    assert summary.refund_amount == Decimal("200000.00")
    assert summary.net_amount == Decimal("1800000.00")
    assert summary.center_net == Decimal("540000.00")
    assert summary.tutor_net == Decimal("1260000.00")

    workbook = load_workbook(BytesIO(build_finance_excel(rows)))
    sheet = workbook["Chi tiet doanh thu"]
    assert sheet.freeze_panes == "A2"
    assert sheet.tables["FinanceReport"].ref == "A1:AG3"
    assert sheet["I2"].value == "'=Học viên kiểm thử"
    assert "'=Điều chỉnh kiểm thử" in {sheet["W2"].value, sheet["W3"].value}


@pytest.mark.asyncio
async def test_finance_roles_and_tutor_scope(
    client: AsyncClient,
    db_session: AsyncSession,
):
    data = await seed_finance_data(db_session)
    global current_test_user_id

    current_test_user_id = data["tutor_account"].id
    response = await client.get("/api/v1/finance/summary")
    assert response.status_code == 403

    response = await client.get("/api/v1/tutor/income/transactions")
    assert response.status_code == 200
    tutor_rows = response.json()["data"]
    assert len(tutor_rows) == 2
    assert all(row["tutor_net"] != "300000.00" for row in tutor_rows)

    current_test_user_id = data["staff"].id
    response = await client.get("/api/v1/finance/summary")
    assert response.status_code == 200

    response = await client.get("/api/v1/finance/export.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert load_workbook(BytesIO(response.content))["Chi tiet doanh thu"].max_row == 4

    response = await client.get("/api/v1/tutor/income/summary")
    assert response.status_code == 403


@pytest.mark.asyncio
async def test_contract_commission_override_and_snapshot_lock(
    client: AsyncClient,
    db_session: AsyncSession,
):
    data = await seed_finance_data(db_session)
    global current_test_user_id
    current_test_user_id = data["staff"].id

    response = await client.put(
        f"/api/v1/contracts/{data['editable_contract'].id}/commission",
        json={"center_rate": 35, "tutor_rate": 65, "reason": "Tỷ lệ hợp đồng riêng"},
    )
    assert response.status_code == 200

    audit_result = await db_session.execute(
        select(AuditLog).where(AuditLog.action == "UPDATE_COMMISSION")
    )
    assert audit_result.scalar_one_or_none() is not None

    response = await client.put(
        f"/api/v1/contracts/{data['contract'].id}/commission",
        json={"center_rate": 35, "tutor_rate": 65, "reason": "Không được phép"},
    )
    assert response.status_code == 400
